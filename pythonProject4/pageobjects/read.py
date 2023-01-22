from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
driver=webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/AutomationPractice/")
wb = openpyxl.Workbook()
sheet=wb.active
sheet.title="mytitle"
number_of_rows = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr")
i=1
for item in number_of_rows:
    sheet.cell(i,1).value=driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[1]").text
    sheet.cell(i,2).value=driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[2]").text
    sheet.cell(i,3).value=driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[3]").text
    sheet.cell(i,4).value=driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[4]").text

    i=i+1
wb.save('C:\\Users\\nikub\\PycharmProjects\\pythonProject4\\Testdata\\demo1.xlsx')

