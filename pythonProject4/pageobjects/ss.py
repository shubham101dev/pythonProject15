import openpyxl
# workbook=openpyxl.load_workbook('C:\\Users\\nikub\\PycharmProjects\\pythonProject4\\Testdata\\Book1.xlsx')
# read data from excel
# sheet=workbook['ss']
# a=sheet['A1'].value
# print(a)
# sheet1=workbook.active
# b=sheet1['B1'].value
# print(b)
# import openpyxl module
# import openpyxl
#
# # Call a Workbook() function of openpyxl
# # to create a new blank Workbook object
wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1).value="shubham"

c2 = sheet.cell(row= 1 , column = 2).value="nikhik"

c3 = sheet['A2']
c3.value = "RAHUL"
c4 = sheet['B2']
c4.value = "RAI"
wb.save("demo.xlsx")
from selenium.webdriver.common.by import By
from selenium import webdriver
wb = openpyxl.Workbook()
sheet=wb.active
sheet.title="mytitle"
driver=webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/AutomationPractice/")
sheet.cell (row=1,column=1).value=driver.find_element(By.XPATH,"//td[normalize-space()='Businessman']").text

wb.save('C:\\Users\\nikub\\PycharmProjects\\pythonProject4\\Testdata\\demo.xlsx')
