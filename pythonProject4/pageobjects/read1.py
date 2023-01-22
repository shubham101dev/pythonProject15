from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
driver=webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/AutomationPractice/")
wb = openpyxl.Workbook()
sheet=wb.active
sheet.title="mytitle"
number_of_rows = driver.find_elements(By.XPATH,"//div[@class='left-align']//tbody/tr")

i=2
for item in range(len(number_of_rows)-1):
    sheet.cell(i,1).value=driver.find_element(By.XPATH,"//div[@class='left-align']//tbody/tr["+str(i)+"]/td[1]").text
    sheet.cell(i,2).value=driver.find_element(By.XPATH,"//div[@class='left-align']//tbody/tr["+str(i)+"]/td[2]").text
    sheet.cell(i,3).value=driver.find_element(By.XPATH,"//div[@class='left-align']//tbody/tr["+str(i)+"]/td[3]").text

    i=i+1
    # s=s+1
wb.save('C:\\Users\\nikub\\PycharmProjects\\pythonProject4\\Testdata\\demo155.xlsx')