import openpyxl


# wb=openpyxl.load_workbook("C:\\Users\\nikub\\PycharmProjects\\pythonProject5\\testdata\\Book6.xlsx")
from selenium import webdriver
from selenium.webdriver.common.by import By

workbook=openpyxl.Workbook()
sheet=workbook.active
sheet.title='shubham'
driver=webdriver.Chrome()
driver.implicitly_wait(5)
driver.get('https://rahulshettyacademy.com/AutomationPractice/')
# a=driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr")
# i=1
# for item in a:
#     sheet.cell(i,1).value=driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[1]").text
#     sheet.cell(i,2).value=driver.find_element(By.XPATH, "//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[2]").text
#     sheet.cell(i,3).value=driver.find_element(By.XPATH, "//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[3]").text
#     sheet.cell(i,4).value=driver.find_element(By.XPATH, "//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[4]").text
#     i=i+1
# workbook.save('shubh5.xlsx')
b=driver.find_elements(By.XPATH,"//table[@id='product' and @name='courses']/tbody/tr")
i=2
j=1
for item in range(len(b)-1):
    sheet.cell(j,1).value=driver.find_element(By.XPATH,"//table[@id='product' and @name='courses']/tbody/tr["+str(i)+"]/td[1]").text
    sheet.cell(j,2).value=driver.find_element(By.XPATH,"//table[@id='product' and @name='courses']/tbody/tr["+str(i)+"]/td[2]").text
    sheet.cell(j,3).value=driver.find_element(By.XPATH,"//table[@id='product' and @name='courses']/tbody/tr["+str(i)+"]/td[3]").text
    i=i+1
    j=j+1
workbook.save('shubh5.xlsx')
driver.quit()


